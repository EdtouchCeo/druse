"""
교사별/학반별 시간표 XLS 파싱 스크립트
출력: output/web/_timetable_data.json
"""
import openpyxl
import json
import sys
import os
import re

DAYS = ['월', '화', '수', '목', '금']
PERIODS = 7
BLOCK_MARKS = ('│', '▽')   # 블록타임(연강) 표시 — 실제 수업/교사가 아님


def norm_teacher_name(name):
    """교사명 정규화. 공동 수업 블록 '한상우/하태희' → '한상우(하태희)'."""
    s = str(name or '').replace('\n', '').strip()
    m = re.match(r'^([^/]+)/(.+)$', s)
    return '%s(%s)' % (m.group(1).strip(), m.group(2).strip()) if m else s

def parse_teacher_schedule(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    total_cols = ws.max_column

    teachers = {}

    # 교사 블록 헤더 위치 탐색 ('교사 시간표' 셀)
    for row_i, row in enumerate(rows):
        for col_i, cell in enumerate(row):
            if cell == '교사 시간표':
                # 교사 이름: 다음 행 col+3
                if row_i + 1 < len(rows) and col_i + 3 < len(rows[row_i + 1]):
                    name = norm_teacher_name(rows[row_i + 1][col_i + 3])
                    if not name:
                        continue

                    # 교시 데이터 파싱: row+3부터 2행씩 7교시
                    schedule = {d: [] for d in DAYS}
                    base = row_i + 3  # 1교시 시작행

                    for period in range(1, PERIODS + 1):
                        subj_row = base + (period - 1) * 2
                        class_row = subj_row + 1
                        if subj_row >= len(rows):
                            break

                        for d_idx, day in enumerate(DAYS):
                            col = col_i + 1 + d_idx
                            if col < len(rows[subj_row]):
                                subj = rows[subj_row][col]
                                cls = rows[class_row][col] if class_row < len(rows) else None
                                schedule[day].append({
                                    'period': period,
                                    'subject': str(subj).strip() if subj else '',
                                    'class': str(cls).strip() if cls else ''
                                })

                    teachers[name] = schedule

    return teachers


def parse_class_schedule(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    classes = {}

    # 학반 블록 헤더 위치 탐색 ('학반 시간표' 셀)
    for row_i, row in enumerate(rows):
        for col_i, cell in enumerate(row):
            if cell == '학반 시간표':
                # 학반 정보: 다음 행 col+3 (예: "1-1 안홍엽")
                if row_i + 1 < len(rows) and col_i + 3 < len(rows[row_i + 1]):
                    class_info = rows[row_i + 1][col_i + 3]
                    if not class_info or not str(class_info).strip():
                        continue
                    class_info = str(class_info).strip()

                    # 학반 이름 (예: "1-1"), 담임 (예: "안홍엽")
                    parts = class_info.split(' ', 1)
                    class_name = parts[0]
                    homeroom = parts[1] if len(parts) > 1 else ''

                    # 교시 데이터 파싱
                    schedule = {d: [] for d in DAYS}
                    base = row_i + 3

                    for period in range(1, PERIODS + 1):
                        subj_row = base + (period - 1) * 2
                        teacher_row = subj_row + 1
                        if subj_row >= len(rows):
                            break

                        for d_idx, day in enumerate(DAYS):
                            col = col_i + 1 + d_idx
                            if col < len(rows[subj_row]):
                                subj = rows[subj_row][col]
                                teacher = rows[teacher_row][col] if teacher_row < len(rows) else None
                                schedule[day].append({
                                    'period': period,
                                    'subject': str(subj).strip() if subj else '',
                                    'teacher': str(teacher).strip() if teacher else ''
                                })

                    classes[class_name] = {
                        'homeroom': homeroom,
                        'schedule': schedule
                    }

    return classes


def teachers_from_classes(classes, known, aliases):
    """학급 시간표를 역인덱싱해 교사별 시간표를 만든다.

    교사별 파일에 개인 블록이 없는 교사(예: 2026-2 이강산)를 학급 시간표 기준으로 복원한다.
    `known`(교사별 파일에서 이미 얻은 교사)이거나 `aliases`로 이미 연결되는 이름은 만들지 않는다.
    """
    derived = {}
    for class_name, info in classes.items():
        for day in DAYS:
            for slot in info['schedule'].get(day, []):
                subject = slot['subject'].strip()
                teacher = norm_teacher_name(slot['teacher'])
                if not subject or not teacher:
                    continue
                if subject.startswith(BLOCK_MARKS) or teacher.startswith(BLOCK_MARKS):
                    continue          # 블록타임 표시(│/▽)는 수업이 아님
                if teacher in known or slot['teacher'].strip() in aliases:
                    continue
                sched = derived.setdefault(
                    teacher,
                    {d: [{'period': p + 1, 'subject': '', 'class': ''} for p in range(PERIODS)]
                     for d in DAYS})
                cell = sched[day][slot['period'] - 1]
                cell['subject'], cell['class'] = subject, class_name
    return derived


def alias_map(teachers, classes):
    """학급 시간표의 교사명 → 교사별 시간표의 키. 표기가 다른 경우만 담는다.

    예: 학급측 '한상우' → 교사키 '한상우(하태희)' (공동 수업 블록)
    """
    aliases = {}
    for info in classes.values():
        for day in DAYS:
            for slot in info['schedule'].get(day, []):
                name = slot['teacher'].strip()
                if not name or name in teachers or name.startswith(BLOCK_MARKS):
                    continue
                hit = [k for k in teachers if k.startswith(name + '(') or k == norm_teacher_name(name)]
                if len(hit) == 1:
                    aliases[name] = hit[0]
    return aliases


def main():
    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    # 기본값 = 2026학년도 2학기. 다른 학기는 인자로 전달: parse_timetable.py <교사별> <학반별>
    src = os.path.join(base, 'input', 'teacher', '2026_2학기_시간표')
    teacher_file = sys.argv[1] if len(sys.argv) > 1 else os.path.join(src, '교사별 시간표(2026-2).xlsx')
    class_file = sys.argv[2] if len(sys.argv) > 2 else os.path.join(src, '학급별 시간표(2026-2).xlsx')
    out_file = os.path.join(base, 'output', 'web', '_timetable_data.json')
    compact_file = os.path.join(base, 'output', 'web', '_timetable_compact.json')
    alias_file = os.path.join(base, 'output', 'web', '_timetable_alias.json')

    print(f"교사별 시간표 파싱: {teacher_file}")
    teachers = parse_teacher_schedule(teacher_file)
    print(f"  → {len(teachers)}명 추출")

    print(f"학반별 시간표 파싱: {class_file}")
    classes = parse_class_schedule(class_file)
    print(f"  → {len(classes)}반 추출")

    # 표기가 다른 교사명을 먼저 연결(예: 학급측 '한상우' → 교사키 '한상우(하태희)').
    # 별칭을 먼저 잡아야 같은 교사가 별도 항목으로 중복 생성되지 않는다.
    aliases = alias_map(teachers, classes)
    print(f"교사명 별칭: {aliases}")

    # 교사별 파일에 개인 블록이 빠진 교사는 학급 시간표를 역인덱싱해 복원
    derived = teachers_from_classes(classes, teachers, aliases)
    for name, schedule in derived.items():
        teachers[name] = schedule
    if derived:
        print(f"학급 시간표 기준 복원: {len(derived)}명 {sorted(derived)}")

    data = {
        'teachers': teachers,
        'classes': classes
    }

    with open(out_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"저장 완료: {out_file}")

    # 웹페이지 삽입용 compact 형식 — 슬롯 1개를 "과목|학반" / "과목|교사" 문자열로
    compact = {
        't': {n: {d: ['%s|%s' % (s['subject'], s['class']) for s in sch[d]] for d in DAYS}
              for n, sch in teachers.items()},
        'c': {n: {'h': v['homeroom'],
                  's': {d: ['%s|%s' % (s['subject'], s['teacher']) for s in v['schedule'][d]] for d in DAYS}}
              for n, v in classes.items()},
    }
    with open(compact_file, 'w', encoding='utf-8') as f:
        json.dump(compact, f, ensure_ascii=False, separators=(',', ':'))
    print(f"저장 완료: {compact_file} ({os.path.getsize(compact_file):,} bytes)")

    with open(alias_file, 'w', encoding='utf-8') as f:
        json.dump(aliases, f, ensure_ascii=False, separators=(',', ':'))
    print(f"저장 완료: {alias_file}")
    return data


if __name__ == '__main__':
    main()
