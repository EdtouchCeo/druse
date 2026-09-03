#!/usr/bin/env python3
"""Build the public round-two club recruitment snapshot from the source workbooks.

Only aggregate, explicitly allowlisted fields are ever copied into the output. The
applicant workbook is opened read-only and is never saved by this script.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl


PHASE = "round2-recruitment-guide"
PHASE_LABEL = "2차 지원 가능 인원 안내"
TERM = "2026학년도 2학기"
OUTPUT_PREFIX = "window.DR_CLUB_STATS_2026_2 = "
CATALOG_SHEET = "1_2학년"
RECRUITMENT_SHEET = "추가 모집 현황"
SUMMARY_SHEET = "학년별 지원 현황"
PREORGANIZED_CLUB = "축구반"
AI_YOUTH_CLUB = "AI유스프러너"
VALID_STATUSES = frozenset({"확정", "기확정", "선발 대기"})
CATALOG_HEADERS = ("연번", "개설 주체", "부서명", "인원")
RECRUITMENT_HEADERS = (
    "개설주체",
    "동아리명",
    "1학년 현재",
    "1학년 기준",
    "1학년 추가모집",
    "2학년 현재",
    "2학년 기준",
    "2학년 추가모집",
    "총 현재",
    "총 기준",
    "총 추가모집",
)


@dataclass(frozen=True)
class CatalogClub:
    catalog_order: int
    name: str
    group: str
    catalog_capacity: int


@dataclass(frozen=True)
class RecruitmentRow:
    order: int
    name: str
    group: str
    grade1_current: int
    grade1_target: int
    grade1_available: int
    grade2_current: int
    grade2_target: int
    grade2_available: int
    total_current: int
    capacity: int
    total_available: int


def normalized_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def integer(value: Any, where: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{where}은 정수여야 합니다.")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    raise ValueError(f"{where}은 정수여야 합니다: {value!r}")


def group_code(value: Any, where: str) -> str:
    label = normalized_text(value)
    try:
        return {"학생": "student", "교사": "teacher"}[label]
    except KeyError as exc:
        raise ValueError(f"{where}의 개설 주체가 올바르지 않습니다: {label!r}") from exc


def operational_capacity(group: str) -> int:
    return 22 if group == "student" else 21


def targets_for(group: str, grade1_current: int, grade2_current: int) -> tuple[int, int]:
    if group == "student":
        return 11, 11
    if grade2_current > grade1_current:
        return 10, 11
    return 11, 10


def open_workbook(path: Path):
    if not path.is_file():
        raise ValueError(f"Excel 파일을 찾을 수 없습니다: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f".xlsx 파일만 지원합니다: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def load_catalog(path: Path) -> tuple[list[CatalogClub], list[str]]:
    workbook = open_workbook(path)
    try:
        if CATALOG_SHEET not in workbook.sheetnames:
            raise ValueError(f"개설 현황에 {CATALOG_SHEET!r} 시트가 없습니다.")
        sheet = workbook[CATALOG_SHEET]
        headers = tuple(normalized_text(sheet.cell(1, column).value) for column in range(1, 5))
        if headers != CATALOG_HEADERS:
            raise ValueError(f"개설 현황 머리글이 올바르지 않습니다: {headers!r}")

        clubs: list[CatalogClub] = []
        diagnostics: list[str] = []
        seen_names: set[str] = set()
        for row in range(2, sheet.max_row + 1):
            raw_values = [sheet.cell(row, column).value for column in range(1, 5)]
            if not any(value is not None for value in raw_values):
                continue
            order = integer(raw_values[0], f"개설 현황 {row}행 연번")
            group = group_code(raw_values[1], f"개설 현황 {row}행")
            name = normalized_text(raw_values[2])
            if not name:
                raise ValueError(f"개설 현황 {row}행의 동아리명이 비어 있습니다.")
            if name in seen_names:
                raise ValueError(f"개설 현황에 동아리명이 중복되었습니다: {name}")
            seen_names.add(name)
            catalog_capacity = integer(raw_values[3], f"개설 현황 {name} 정원")
            expected_capacity = operational_capacity(group)
            if catalog_capacity != expected_capacity:
                if name == AI_YOUTH_CLUB and group == "student" and catalog_capacity == 21:
                    diagnostics.append(
                        f"개설 현황의 {AI_YOUTH_CLUB} 정원 21명은 운영 기준 22명으로 보정했습니다."
                    )
                else:
                    raise ValueError(
                        f"개설 현황 정원 불일치: {name}={catalog_capacity}명 "
                        f"(운영 기준 {expected_capacity}명)"
                    )
            clubs.append(CatalogClub(order, name, group, catalog_capacity))

        if len(clubs) != 29:
            raise ValueError(f"개설 현황의 공식 동아리는 29개여야 합니다: {len(clubs)}개")
        orders = [club.catalog_order for club in clubs]
        if orders != list(range(1, 30)):
            raise ValueError(f"개설 현황 연번은 1~29 순서여야 합니다: {orders!r}")
        return clubs, diagnostics
    finally:
        workbook.close()


def count_current_by_club(path: Path, clubs: list[CatalogClub]) -> tuple[dict[str, tuple[int, int]], int, int]:
    workbook = open_workbook(path)
    try:
        missing = [club.name for club in clubs if club.name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"지원 현황에 공식 동아리 시트가 없습니다: {missing!r}")

        counts: dict[str, tuple[int, int]] = {}
        valid_rows = 0
        rows_without_valid_status = 0
        for club in clubs:
            sheet = workbook[club.name]
            if normalized_text(sheet.cell(1, 1).value) != "학년" or normalized_text(sheet.cell(1, 8).value) != "배정 상태":
                raise ValueError(f"{club.name} 시트의 학년/배정 상태 머리글이 올바르지 않습니다.")
            grade_counts = Counter({1: 0, 2: 0})
            for row in range(2, sheet.max_row + 1):
                status = normalized_text(sheet.cell(row, 8).value)
                row_has_data = any(sheet.cell(row, column).value is not None for column in range(1, 9))
                if status in VALID_STATUSES:
                    grade = integer(sheet.cell(row, 1).value, f"{club.name} {row}행 학년")
                    if grade not in (1, 2):
                        raise ValueError(f"{club.name} {row}행 학년은 1 또는 2여야 합니다.")
                    grade_counts[grade] += 1
                    valid_rows += 1
                elif row_has_data:
                    rows_without_valid_status += 1
            counts[club.name] = (grade_counts[1], grade_counts[2])
        return counts, valid_rows, rows_without_valid_status
    finally:
        workbook.close()


def computed_recruitment(
    clubs: list[CatalogClub], counts: dict[str, tuple[int, int]]
) -> dict[str, RecruitmentRow]:
    result: dict[str, RecruitmentRow] = {}
    for club in clubs:
        if club.name == PREORGANIZED_CLUB:
            continue
        grade1_current, grade2_current = counts[club.name]
        grade1_target, grade2_target = targets_for(club.group, grade1_current, grade2_current)
        capacity = operational_capacity(club.group)
        if grade1_target + grade2_target != capacity:
            raise ValueError(f"{club.name}의 학년별 기준 합이 운영 정원과 다릅니다.")
        grade1_available = max(grade1_target - grade1_current, 0)
        grade2_available = max(grade2_target - grade2_current, 0)
        total_available = grade1_available + grade2_available
        if total_available == 0:
            continue
        result[club.name] = RecruitmentRow(
            order=0,
            name=club.name,
            group=club.group,
            grade1_current=grade1_current,
            grade1_target=grade1_target,
            grade1_available=grade1_available,
            grade2_current=grade2_current,
            grade2_target=grade2_target,
            grade2_available=grade2_available,
            total_current=grade1_current + grade2_current,
            capacity=capacity,
            total_available=total_available,
        )
    return result


def load_recruitment_table(path: Path) -> tuple[str, list[RecruitmentRow]]:
    workbook = open_workbook(path)
    try:
        if RECRUITMENT_SHEET not in workbook.sheetnames:
            raise ValueError(f"지원 현황에 {RECRUITMENT_SHEET!r} 시트가 없습니다.")
        sheet = workbook[RECRUITMENT_SHEET]
        if normalized_text(sheet.cell(1, 1).value) != RECRUITMENT_SHEET:
            raise ValueError("추가 모집 현황 제목이 올바르지 않습니다.")
        basis_match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", normalized_text(sheet.cell(2, 1).value))
        if not basis_match:
            raise ValueError("추가 모집 현황에서 기준일을 찾을 수 없습니다.")
        basis_date = basis_match.group(1)
        try:
            date.fromisoformat(basis_date)
        except ValueError as exc:
            raise ValueError(f"추가 모집 기준일이 올바르지 않습니다: {basis_date}") from exc

        headers = tuple(normalized_text(sheet.cell(3, column).value) for column in range(1, 12))
        if headers != RECRUITMENT_HEADERS:
            raise ValueError(f"추가 모집 현황 머리글이 올바르지 않습니다: {headers!r}")

        rows: list[RecruitmentRow] = []
        total_row: list[Any] | None = None
        for row in range(4, sheet.max_row + 1):
            first = normalized_text(sheet.cell(row, 1).value)
            if first == "합계":
                total_row = [sheet.cell(row, column).value for column in range(1, 12)]
                break
            raw_values = [sheet.cell(row, column).value for column in range(1, 12)]
            if not any(value is not None for value in raw_values):
                continue
            name = normalized_text(raw_values[1])
            if not name:
                raise ValueError(f"추가 모집 현황 {row}행의 동아리명이 비어 있습니다.")
            values = [
                integer(raw_values[index], f"추가 모집 현황 {name} {RECRUITMENT_HEADERS[index]}")
                for index in range(2, 11)
            ]
            rows.append(
                RecruitmentRow(
                    order=len(rows) + 1,
                    name=name,
                    group=group_code(raw_values[0], f"추가 모집 현황 {name}"),
                    grade1_current=values[0],
                    grade1_target=values[1],
                    grade1_available=values[2],
                    grade2_current=values[3],
                    grade2_target=values[4],
                    grade2_available=values[5],
                    total_current=values[6],
                    capacity=values[7],
                    total_available=values[8],
                )
            )

        if not rows or total_row is None:
            raise ValueError("추가 모집 현황의 데이터 행 또는 합계 행을 찾을 수 없습니다.")
        names = [row.name for row in rows]
        if len(names) != len(set(names)):
            raise ValueError("추가 모집 현황에 동아리명이 중복되었습니다.")

        expected_totals = [
            sum(row.grade1_current for row in rows),
            sum(row.grade1_target for row in rows),
            sum(row.grade1_available for row in rows),
            sum(row.grade2_current for row in rows),
            sum(row.grade2_target for row in rows),
            sum(row.grade2_available for row in rows),
            sum(row.total_current for row in rows),
            sum(row.capacity for row in rows),
            sum(row.total_available for row in rows),
        ]
        actual_totals = [
            integer(total_row[index], f"추가 모집 합계 {RECRUITMENT_HEADERS[index]}")
            for index in range(2, 11)
        ]
        if actual_totals != expected_totals:
            raise ValueError(f"추가 모집 현황 합계 행이 데이터 행 합계와 다릅니다: {actual_totals!r}")
        if normalized_text(total_row[1]) != f"{len(rows)}개 동아리":
            raise ValueError("추가 모집 현황 합계 행의 동아리 수가 올바르지 않습니다.")
        return basis_date, rows
    finally:
        workbook.close()


def cross_validate(
    computed: dict[str, RecruitmentRow], table_rows: list[RecruitmentRow]
) -> list[RecruitmentRow]:
    computed_names = set(computed)
    table_names = {row.name for row in table_rows}
    if computed_names != table_names:
        missing = sorted(computed_names - table_names)
        unexpected = sorted(table_names - computed_names)
        raise ValueError(f"추가 모집 대상 동아리 집합 불일치: 누락={missing!r}, 초과={unexpected!r}")

    validated: list[RecruitmentRow] = []
    comparable_fields = (
        "group",
        "grade1_current",
        "grade1_target",
        "grade1_available",
        "grade2_current",
        "grade2_target",
        "grade2_available",
        "total_current",
        "capacity",
        "total_available",
    )
    for table_row in table_rows:
        calculated = computed[table_row.name]
        differences = [
            field for field in comparable_fields if getattr(table_row, field) != getattr(calculated, field)
        ]
        if differences:
            raise ValueError(f"추가 모집 현황과 재계산 값 불일치: {table_row.name} {differences!r}")
        validated.append(
            RecruitmentRow(
                order=table_row.order,
                name=calculated.name,
                group=calculated.group,
                grade1_current=calculated.grade1_current,
                grade1_target=calculated.grade1_target,
                grade1_available=calculated.grade1_available,
                grade2_current=calculated.grade2_current,
                grade2_target=calculated.grade2_target,
                grade2_available=calculated.grade2_available,
                total_current=calculated.total_current,
                capacity=calculated.capacity,
                total_available=calculated.total_available,
            )
        )
    return validated


def summary_diagnostics(path: Path, counts: dict[str, tuple[int, int]], clubs: list[CatalogClub]) -> list[str]:
    """Compare only the aggregate part (rows 4-33), never the PII rows below it."""
    workbook = open_workbook(path)
    try:
        if SUMMARY_SHEET not in workbook.sheetnames:
            return [f"{SUMMARY_SHEET!r} 시트가 없어 진단을 건너뜁니다."]
        sheet = workbook[SUMMARY_SHEET]
        diagnostics: list[str] = []
        catalog_names = {club.name for club in clubs}
        for row in range(4, min(sheet.max_row, 33) + 1):
            name = normalized_text(sheet.cell(row, 1).value)
            if name == "합계":
                break
            if name not in catalog_names:
                continue
            values = (sheet.cell(row, 2).value, sheet.cell(row, 3).value)
            if not all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values):
                continue
            summary_counts = (int(values[0]), int(values[1]))
            if summary_counts != counts[name]:
                diagnostics.append(
                    f"학년별 지원 현황 진단: {name} 요약 {summary_counts[0]}/{summary_counts[1]}, "
                    f"유효 배정 행 {counts[name][0]}/{counts[name][1]}"
                )
        return diagnostics
    finally:
        workbook.close()


def build_public(basis_date: str, rows: list[RecruitmentRow], phase: str) -> dict[str, Any]:
    if phase != PHASE:
        raise ValueError(f"지원하지 않는 단계입니다: {phase}")
    clubs = []
    for row in rows:
        clubs.append(
            {
                "order": row.order,
                "name": row.name,
                "group": row.group,
                "groupLabel": "학생 주도" if row.group == "student" else "교사 주도",
                "capacity": row.capacity,
                "grade1": {
                    "current": row.grade1_current,
                    "target": row.grade1_target,
                    "available": row.grade1_available,
                },
                "grade2": {
                    "current": row.grade2_current,
                    "target": row.grade2_target,
                    "available": row.grade2_available,
                },
                "totalAvailable": row.total_available,
            }
        )

    return {
        "schemaVersion": 2,
        "term": TERM,
        "phase": phase,
        "phaseLabel": PHASE_LABEL,
        "basisDate": basis_date,
        "totals": {
            "clubs": len(clubs),
            "groups": {
                "student": sum(club["group"] == "student" for club in clubs),
                "teacher": sum(club["group"] == "teacher" for club in clubs),
            },
            "grade1": {
                "clubs": sum(club["grade1"]["available"] > 0 for club in clubs),
                "available": sum(club["grade1"]["available"] for club in clubs),
            },
            "grade2": {
                "clubs": sum(club["grade2"]["available"] > 0 for club in clubs),
                "available": sum(club["grade2"]["available"] for club in clubs),
            },
            "available": sum(club["totalAvailable"] for club in clubs),
        },
        "rules": {
            "student": {"capacity": 22, "grade1Target": 11, "grade2Target": 11},
            "teacher": {
                "capacity": 21,
                "largerGradeTarget": 11,
                "otherGradeTarget": 10,
                "tieGets11": 1,
            },
            "formula": "max(target-current,0)",
        },
        "clubs": clubs,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True, help="지원 현황 기준본 .xlsx")
    parser.add_argument("--catalog", type=Path, required=True, help="공식 개설 현황 .xlsx")
    parser.add_argument("--output", type=Path, required=True, help="공개 JS 출력 경로")
    parser.add_argument("--phase", default=PHASE)
    args = parser.parse_args()

    catalog, catalog_diagnostics = load_catalog(args.catalog)
    counts, valid_rows, rows_without_valid_status = count_current_by_club(args.workbook, catalog)
    calculated = computed_recruitment(catalog, counts)
    basis_date, table_rows = load_recruitment_table(args.workbook)
    validated_rows = cross_validate(calculated, table_rows)
    public = build_public(basis_date, validated_rows, args.phase)
    diagnostics = catalog_diagnostics + summary_diagnostics(args.workbook, counts, catalog)

    rendered = OUTPUT_PREFIX + json.dumps(public, ensure_ascii=False, indent=2) + ";\n"
    args.output.parent.mkdir(parents=True, exist_ok=True)
    temporary = args.output.with_suffix(args.output.suffix + ".tmp")
    temporary.write_text(rendered, encoding="utf-8", newline="\n")
    temporary.replace(args.output)

    for message in diagnostics:
        print(f"진단: {message}", file=sys.stderr)
    print(
        f"공개 통계 스냅샷 생성 완료: {len(public['clubs'])}개 동아리, "
        f"1학년 {public['totals']['grade1']['available']}명, "
        f"2학년 {public['totals']['grade2']['available']}명, "
        f"전체 {public['totals']['available']}명 "
        f"(유효 배정 행 {valid_rows}건, 상태 없는 내부 행 {rows_without_valid_status}건)"
    )


if __name__ == "__main__":
    main()
