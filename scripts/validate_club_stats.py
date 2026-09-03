#!/usr/bin/env python3
"""Validate the round-two public club snapshot and its source boundaries."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl


PREFIX = "window.DR_CLUB_STATS_2026_2 = "
TERM = "2026학년도 2학기"
PHASE = "round2-recruitment-guide"
PHASE_LABEL = "2차 지원 가능 인원 안내"
CATALOG_SHEET = "1_2학년"
RECRUITMENT_SHEET = "추가 모집 현황"
PREORGANIZED_CLUB = "축구반"
AI_YOUTH_CLUB = "AI유스프러너"
VALID_STATUSES = frozenset({"확정", "기확정", "선발 대기"})

ROOT_KEYS = {"schemaVersion", "term", "phase", "phaseLabel", "basisDate", "totals", "rules", "clubs"}
TOTAL_KEYS = {"clubs", "groups", "grade1", "grade2", "available"}
GROUP_TOTAL_KEYS = {"student", "teacher"}
GRADE_TOTAL_KEYS = {"clubs", "available"}
RULE_KEYS = {"student", "teacher", "formula"}
STUDENT_RULE_KEYS = {"capacity", "grade1Target", "grade2Target"}
TEACHER_RULE_KEYS = {"capacity", "largerGradeTarget", "otherGradeTarget", "tieGets11"}
CLUB_KEYS = {"order", "name", "group", "groupLabel", "capacity", "grade1", "grade2", "totalAvailable"}
CLUB_GRADE_KEYS = {"current", "target", "available"}

FORBIDDEN_KEYS = {
    "privacyNotice",
    "advisor",
    "advisorKey",
    "location",
    "requestedLocation",
    "assignmentTeacher",
    "leader",
    "viceLeader",
    "description",
    "studentIntroduction",
    "selection",
    "sessions",
    "applicationUrl",
    "applicantWorkbookUrl",
    "sources",
    "source",
    "sheet",
    "status",
    "studentName",
    "classNumber",
    "motivation",
    "activityPlan",
}
FORBIDDEN_TEXT = (
    "sharepoint.com",
    "forms.cloud.microsoft",
    "onedrive",
    "@",
    ".xlsx",
    ".xls",
    ".hwpx",
    ".pdf",
)
CATALOG_HEADERS = ("연번", "개설 주체", "부서명", "인원")
RECRUITMENT_HEADERS = (
    "개설주체",
    "동아리명",
    "1학년 현재",
    "1학년 기준",
    "1학년 추가모집",
    "2학년 현재",
    "2학년 기준",
    "2학년 추가모집",
    "총 현재",
    "총 기준",
    "총 추가모집",
)


def normalized_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def source_integer(value: Any, where: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{where}은 정수여야 합니다.")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    raise ValueError(f"{where}은 정수여야 합니다.")


def public_integer(value: Any, where: str, minimum: int = 0) -> int:
    if type(value) is not int or value < minimum:
        raise ValueError(f"{where}은 {minimum} 이상의 정수여야 합니다.")
    return value


def exact_keys(value: Any, expected: set[str], where: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError(f"{where}은 객체여야 합니다.")
    actual = set(value)
    if actual != expected:
        raise ValueError(
            f"{where} 키 불일치: 누락={sorted(expected - actual)!r}, "
            f"허용되지 않음={sorted(actual - expected)!r}"
        )
    return value


def nonempty_string(value: Any, where: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{where}은 비어 있지 않은 문자열이어야 합니다.")
    return value


def load_public(path: Path) -> tuple[dict[str, Any], str]:
    text = path.read_text(encoding="utf-8").strip()
    if not text.startswith(PREFIX) or not text.endswith(";"):
        raise ValueError("공개 JS 래퍼 형식이 올바르지 않습니다.")
    payload = text[len(PREFIX) : -1].strip()
    data = json.loads(payload)
    if not isinstance(data, dict):
        raise ValueError("공개 JS 최상위 값은 객체여야 합니다.")
    return data, text


def walk_forbidden_keys(value: Any, where: str = "root") -> None:
    if isinstance(value, dict):
        for key, child in value.items():
            if key in FORBIDDEN_KEYS:
                raise ValueError(f"금지 키 발견: {where}.{key}")
            walk_forbidden_keys(child, f"{where}.{key}")
    elif isinstance(value, list):
        for index, child in enumerate(value):
            walk_forbidden_keys(child, f"{where}[{index}]")


def validate_privacy_markers(data: dict[str, Any], raw: str) -> None:
    walk_forbidden_keys(data)
    lowered = raw.lower()
    for marker in FORBIDDEN_TEXT:
        if marker in lowered:
            raise ValueError(f"금지 문자열 발견: {marker}")
    if re.search(r"(?:[A-Za-z]:[\\/]|(?:file|https?)://|\\\\[^\\\s]+\\)", raw, re.I):
        raise ValueError("로컬/네트워크 경로 또는 URL이 공개 데이터에 포함되어 있습니다.")


def expected_targets(group: str, grade1_current: int, grade2_current: int) -> tuple[int, int]:
    if group == "student":
        return 11, 11
    return (10, 11) if grade2_current > grade1_current else (11, 10)


def validate_public(data: dict[str, Any], raw: str, expect_current_baseline: bool) -> None:
    validate_privacy_markers(data, raw)
    exact_keys(data, ROOT_KEYS, "root")
    if data["schemaVersion"] != 2:
        raise ValueError("schemaVersion은 2여야 합니다.")
    if data["term"] != TERM or data["phase"] != PHASE or data["phaseLabel"] != PHASE_LABEL:
        raise ValueError("학기 또는 공개 단계 메타데이터가 올바르지 않습니다.")
    basis_date = nonempty_string(data["basisDate"], "basisDate")
    if not re.fullmatch(r"20\d{2}-\d{2}-\d{2}", basis_date):
        raise ValueError("basisDate는 YYYY-MM-DD 형식이어야 합니다.")
    try:
        date.fromisoformat(basis_date)
    except ValueError as exc:
        raise ValueError("basisDate가 유효한 날짜가 아닙니다.") from exc

    totals = exact_keys(data["totals"], TOTAL_KEYS, "totals")
    group_totals = exact_keys(totals["groups"], GROUP_TOTAL_KEYS, "totals.groups")
    grade1_totals = exact_keys(totals["grade1"], GRADE_TOTAL_KEYS, "totals.grade1")
    grade2_totals = exact_keys(totals["grade2"], GRADE_TOTAL_KEYS, "totals.grade2")
    for key in ("clubs", "available"):
        public_integer(totals[key], f"totals.{key}")
    for key in GROUP_TOTAL_KEYS:
        public_integer(group_totals[key], f"totals.groups.{key}")
    for label, values in (("grade1", grade1_totals), ("grade2", grade2_totals)):
        for key in GRADE_TOTAL_KEYS:
            public_integer(values[key], f"totals.{label}.{key}")

    rules = exact_keys(data["rules"], RULE_KEYS, "rules")
    student_rules = exact_keys(rules["student"], STUDENT_RULE_KEYS, "rules.student")
    teacher_rules = exact_keys(rules["teacher"], TEACHER_RULE_KEYS, "rules.teacher")
    if student_rules != {"capacity": 22, "grade1Target": 11, "grade2Target": 11}:
        raise ValueError("학생 주도 운영 기준은 22명·11/11이어야 합니다.")
    if teacher_rules != {
        "capacity": 21,
        "largerGradeTarget": 11,
        "otherGradeTarget": 10,
        "tieGets11": 1,
    }:
        raise ValueError("교사 주도 운영 기준은 21명·11/10, 동률 시 1학년 11명이어야 합니다.")
    if rules["formula"] != "max(target-current,0)":
        raise ValueError("추가 모집 계산식이 올바르지 않습니다.")

    clubs = data["clubs"]
    if not isinstance(clubs, list) or not clubs:
        raise ValueError("clubs는 비어 있지 않은 배열이어야 합니다.")
    seen_names: set[str] = set()
    seen_orders: set[int] = set()
    for index, club_value in enumerate(clubs):
        club = exact_keys(club_value, CLUB_KEYS, f"clubs[{index}]")
        order = public_integer(club["order"], f"clubs[{index}].order", minimum=1)
        name = nonempty_string(club["name"], f"clubs[{index}].name")
        if name in seen_names or order in seen_orders:
            raise ValueError("공개 동아리명 또는 순서가 중복되었습니다.")
        seen_names.add(name)
        seen_orders.add(order)
        group = club["group"]
        if group not in ("student", "teacher"):
            raise ValueError(f"{name}의 group이 올바르지 않습니다.")
        expected_label = "학생 주도" if group == "student" else "교사 주도"
        if club["groupLabel"] != expected_label:
            raise ValueError(f"{name}의 groupLabel이 올바르지 않습니다.")
        expected_capacity = 22 if group == "student" else 21
        if public_integer(club["capacity"], f"{name}.capacity") != expected_capacity:
            raise ValueError(f"{name}의 운영 정원이 개설 주체 기준과 다릅니다.")

        grade1 = exact_keys(club["grade1"], CLUB_GRADE_KEYS, f"{name}.grade1")
        grade2 = exact_keys(club["grade2"], CLUB_GRADE_KEYS, f"{name}.grade2")
        for grade_label, grade in (("grade1", grade1), ("grade2", grade2)):
            for key in CLUB_GRADE_KEYS:
                public_integer(grade[key], f"{name}.{grade_label}.{key}")
        expected_grade1_target, expected_grade2_target = expected_targets(
            group, grade1["current"], grade2["current"]
        )
        if (grade1["target"], grade2["target"]) != (
            expected_grade1_target,
            expected_grade2_target,
        ):
            raise ValueError(f"{name}의 학년별 기준이 운영 규칙과 다릅니다.")
        if grade1["target"] + grade2["target"] != expected_capacity:
            raise ValueError(f"{name}의 학년별 기준 합이 운영 정원과 다릅니다.")
        expected_grade1_available = max(grade1["target"] - grade1["current"], 0)
        expected_grade2_available = max(grade2["target"] - grade2["current"], 0)
        if grade1["available"] != expected_grade1_available or grade2["available"] != expected_grade2_available:
            raise ValueError(f"{name}의 학년별 추가 모집 계산이 올바르지 않습니다.")
        expected_total = expected_grade1_available + expected_grade2_available
        if club["totalAvailable"] != expected_total or expected_total <= 0:
            raise ValueError(f"{name}의 총 추가 모집 또는 공개 포함 조건이 올바르지 않습니다.")

    if seen_orders != set(range(1, len(clubs) + 1)):
        raise ValueError("동아리 order는 1부터 이어지는 순서여야 합니다.")

    derived = {
        "clubs": len(clubs),
        "groups": {
            "student": sum(club["group"] == "student" for club in clubs),
            "teacher": sum(club["group"] == "teacher" for club in clubs),
        },
        "grade1": {
            "clubs": sum(club["grade1"]["available"] > 0 for club in clubs),
            "available": sum(club["grade1"]["available"] for club in clubs),
        },
        "grade2": {
            "clubs": sum(club["grade2"]["available"] > 0 for club in clubs),
            "available": sum(club["grade2"]["available"] for club in clubs),
        },
        "available": sum(club["totalAvailable"] for club in clubs),
    }
    if totals != derived:
        raise ValueError(f"공개 합계가 동아리별 재계산 값과 다릅니다: {derived!r}")

    if expect_current_baseline:
        current_expected = {
            "clubs": 18,
            "groups": {"student": 12, "teacher": 6},
            "grade1": {"clubs": 17, "available": 94},
            "grade2": {"clubs": 13, "available": 80},
            "available": 174,
        }
        if basis_date != "2026-09-01" or totals != current_expected:
            raise ValueError("현재 배포 기준선(2026-09-01, 18/94/80/174)과 다릅니다.")
        ai = next((club for club in clubs if club["name"] == AI_YOUTH_CLUB), None)
        if not ai:
            raise ValueError(f"현재 기준선에 {AI_YOUTH_CLUB}가 없습니다.")
        if not (
            ai["group"] == "student"
            and ai["capacity"] == 22
            and ai["grade1"] == {"current": 9, "target": 11, "available": 2}
            and ai["grade2"] == {"current": 11, "target": 11, "available": 0}
            and ai["totalAvailable"] == 2
        ):
            raise ValueError(f"현재 기준선의 {AI_YOUTH_CLUB} 회귀값이 올바르지 않습니다.")


def open_source(path: Path):
    if not path.is_file() or path.suffix.lower() != ".xlsx":
        raise ValueError("원천 경로는 존재하는 .xlsx 파일이어야 합니다.")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def public_string_values(value: Any):
    """Yield exact string values, excluding object keys and serialized substrings."""
    if isinstance(value, str):
        yield value.strip()
    elif isinstance(value, dict):
        for child in value.values():
            yield from public_string_values(child)
    elif isinstance(value, list):
        for child in value:
            yield from public_string_values(child)


def exact_identity_values(values: list[str]) -> set[str]:
    """Return exact source cells plus clearly delimited Korean personal names."""
    identities: set[str] = set()
    for raw_value in values:
        value = normalized_text(raw_value)
        if not value:
            continue
        identities.add(value)
        for token in re.split(r"[,/·&\n]+", value):
            token = token.strip()
            if re.fullmatch(r"[가-힣]{2,5}", token):
                identities.add(token)
    return identities


def source_group(value: Any, where: str) -> str:
    label = normalized_text(value)
    if label == "학생":
        return "student"
    if label == "교사":
        return "teacher"
    raise ValueError(f"{where}의 개설 주체가 올바르지 않습니다.")


def audit_catalog(path: Path) -> tuple[list[tuple[int, str, str, int]], list[str]]:
    workbook = open_source(path)
    try:
        if CATALOG_SHEET not in workbook.sheetnames:
            raise ValueError("개설 현황의 1_2학년 시트가 없습니다.")
        sheet = workbook[CATALOG_SHEET]
        headers = tuple(normalized_text(sheet.cell(1, column).value) for column in range(1, 5))
        if headers != CATALOG_HEADERS:
            raise ValueError("개설 현황 머리글이 올바르지 않습니다.")
        rows: list[tuple[int, str, str, int]] = []
        teacher_names: list[str] = []
        for row in range(2, sheet.max_row + 1):
            values = [sheet.cell(row, column).value for column in range(1, 6)]
            if not any(value is not None for value in values[:4]):
                continue
            order = source_integer(values[0], f"개설 현황 {row}행 연번")
            group = source_group(values[1], f"개설 현황 {row}행")
            name = normalized_text(values[2])
            capacity = source_integer(values[3], f"개설 현황 {row}행 정원")
            if not name:
                raise ValueError("개설 현황 동아리명이 비어 있습니다.")
            expected_capacity = 22 if group == "student" else 21
            if capacity != expected_capacity and not (
                name == AI_YOUTH_CLUB and group == "student" and capacity == 21
            ):
                raise ValueError(f"개설 현황의 {name} 정원이 운영 기준과 다릅니다.")
            teacher = normalized_text(values[4])
            if teacher:
                teacher_names.append(teacher)
            rows.append((order, name, group, capacity))
        if len(rows) != 29 or [row[0] for row in rows] != list(range(1, 30)):
            raise ValueError("개설 현황은 연번 1~29의 공식 동아리 29개여야 합니다.")
        if len({row[1] for row in rows}) != 29:
            raise ValueError("개설 현황에 동아리명이 중복되었습니다.")
        return rows, teacher_names
    finally:
        workbook.close()


def audit_individual_sheets(
    path: Path, catalog: list[tuple[int, str, str, int]]
) -> tuple[dict[str, tuple[int, int]], list[str], list[str]]:
    workbook = open_source(path)
    try:
        missing = [name for _, name, _, _ in catalog if name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"지원 현황에 공식 동아리 시트가 없습니다: {missing!r}")
        counts: dict[str, tuple[int, int]] = {}
        student_names: list[str] = []
        narrative_values: list[str] = []
        for _, name, _, _ in catalog:
            sheet = workbook[name]
            if normalized_text(sheet.cell(1, 1).value) != "학년" or normalized_text(sheet.cell(1, 8).value) != "배정 상태":
                raise ValueError(f"{name} 시트의 필수 머리글이 올바르지 않습니다.")
            grade_counts = Counter({1: 0, 2: 0})
            for row in range(2, sheet.max_row + 1):
                status = normalized_text(sheet.cell(row, 8).value)
                if status in VALID_STATUSES:
                    grade = source_integer(sheet.cell(row, 1).value, f"{name} {row}행 학년")
                    if grade not in (1, 2):
                        raise ValueError(f"{name}의 유효 배정 행 학년이 올바르지 않습니다.")
                    grade_counts[grade] += 1
                student_name = normalized_text(sheet.cell(row, 4).value)
                if student_name:
                    student_names.append(student_name)
                for column in (6, 7):
                    narrative = normalized_text(sheet.cell(row, column).value)
                    if len(narrative) >= 20:
                        narrative_values.append(narrative)
            counts[name] = (grade_counts[1], grade_counts[2])
        return counts, student_names, narrative_values
    finally:
        workbook.close()


def independently_compute(
    catalog: list[tuple[int, str, str, int]], counts: dict[str, tuple[int, int]]
) -> dict[str, dict[str, Any]]:
    computed: dict[str, dict[str, Any]] = {}
    for _, name, group, _ in catalog:
        if name == PREORGANIZED_CLUB:
            continue
        grade1_current, grade2_current = counts[name]
        grade1_target, grade2_target = expected_targets(group, grade1_current, grade2_current)
        grade1_available = max(grade1_target - grade1_current, 0)
        grade2_available = max(grade2_target - grade2_current, 0)
        if grade1_available + grade2_available == 0:
            continue
        computed[name] = {
            "group": group,
            "capacity": 22 if group == "student" else 21,
            "grade1": {
                "current": grade1_current,
                "target": grade1_target,
                "available": grade1_available,
            },
            "grade2": {
                "current": grade2_current,
                "target": grade2_target,
                "available": grade2_available,
            },
            "totalAvailable": grade1_available + grade2_available,
        }
    return computed


def audit_recruitment_table(path: Path) -> tuple[str, list[dict[str, Any]]]:
    workbook = open_source(path)
    try:
        if RECRUITMENT_SHEET not in workbook.sheetnames:
            raise ValueError("지원 현황의 추가 모집 현황 시트가 없습니다.")
        sheet = workbook[RECRUITMENT_SHEET]
        if normalized_text(sheet.cell(1, 1).value) != RECRUITMENT_SHEET:
            raise ValueError("추가 모집 현황 제목이 올바르지 않습니다.")
        match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", normalized_text(sheet.cell(2, 1).value))
        if not match:
            raise ValueError("추가 모집 기준일을 찾을 수 없습니다.")
        basis_date = match.group(1)
        headers = tuple(normalized_text(sheet.cell(3, column).value) for column in range(1, 12))
        if headers != RECRUITMENT_HEADERS:
            raise ValueError("추가 모집 현황 머리글이 올바르지 않습니다.")

        rows: list[dict[str, Any]] = []
        total_values: list[Any] | None = None
        for row in range(4, sheet.max_row + 1):
            first = normalized_text(sheet.cell(row, 1).value)
            if first == "합계":
                total_values = [sheet.cell(row, column).value for column in range(1, 12)]
                break
            values = [sheet.cell(row, column).value for column in range(1, 12)]
            if not any(value is not None for value in values):
                continue
            name = normalized_text(values[1])
            numeric = [source_integer(values[index], f"추가 모집 {name} 값") for index in range(2, 11)]
            row_value = {
                "order": len(rows) + 1,
                "name": name,
                "group": source_group(values[0], f"추가 모집 {name}"),
                "capacity": numeric[7],
                "grade1": {"current": numeric[0], "target": numeric[1], "available": numeric[2]},
                "grade2": {"current": numeric[3], "target": numeric[4], "available": numeric[5]},
                "totalAvailable": numeric[8],
                "totalCurrent": numeric[6],
            }
            if row_value["totalCurrent"] != numeric[0] + numeric[3]:
                raise ValueError(f"추가 모집 {name}의 총 현재가 올바르지 않습니다.")
            if row_value["capacity"] != numeric[1] + numeric[4]:
                raise ValueError(f"추가 모집 {name}의 총 기준이 올바르지 않습니다.")
            if row_value["totalAvailable"] != numeric[2] + numeric[5]:
                raise ValueError(f"추가 모집 {name}의 총 추가모집이 올바르지 않습니다.")
            rows.append(row_value)
        if not rows or total_values is None:
            raise ValueError("추가 모집 현황의 데이터 또는 합계 행이 없습니다.")
        if len({row["name"] for row in rows}) != len(rows):
            raise ValueError("추가 모집 현황에 동아리명이 중복되었습니다.")
        expected_totals = [
            sum(row["grade1"]["current"] for row in rows),
            sum(row["grade1"]["target"] for row in rows),
            sum(row["grade1"]["available"] for row in rows),
            sum(row["grade2"]["current"] for row in rows),
            sum(row["grade2"]["target"] for row in rows),
            sum(row["grade2"]["available"] for row in rows),
            sum(row["totalCurrent"] for row in rows),
            sum(row["capacity"] for row in rows),
            sum(row["totalAvailable"] for row in rows),
        ]
        actual_totals = [source_integer(total_values[index], "추가 모집 합계") for index in range(2, 11)]
        if actual_totals != expected_totals:
            raise ValueError("추가 모집 현황 합계가 데이터 행 합계와 다릅니다.")
        return basis_date, rows
    finally:
        workbook.close()


def validate_sources(
    data: dict[str, Any], workbook_path: Path, catalog_path: Path
) -> None:
    catalog, teacher_names = audit_catalog(catalog_path)
    counts, student_names, narrative_values = audit_individual_sheets(workbook_path, catalog)
    computed = independently_compute(catalog, counts)
    basis_date, table_rows = audit_recruitment_table(workbook_path)
    table_by_name = {row["name"]: row for row in table_rows}
    if set(computed) != set(table_by_name):
        raise ValueError("개별 시트 재계산 결과와 추가 모집 현황의 동아리 집합이 다릅니다.")
    for name, expected in computed.items():
        table = table_by_name[name]
        comparable = {key: table[key] for key in ("group", "capacity", "grade1", "grade2", "totalAvailable")}
        if comparable != expected:
            raise ValueError(f"개별 시트 재계산 값과 추가 모집 현황이 다릅니다: {name}")

    if data["basisDate"] != basis_date:
        raise ValueError("공개 기준일이 추가 모집 현황의 기준일과 다릅니다.")
    public_clubs = data["clubs"]
    if [club["name"] for club in public_clubs] != [row["name"] for row in table_rows]:
        raise ValueError("공개 동아리 순서가 추가 모집 현황의 순서와 다릅니다.")
    for public, table in zip(public_clubs, table_rows, strict=True):
        expected_public = {
            "order": table["order"],
            "name": table["name"],
            "group": table["group"],
            "groupLabel": "학생 주도" if table["group"] == "student" else "교사 주도",
            "capacity": table["capacity"],
            "grade1": table["grade1"],
            "grade2": table["grade2"],
            "totalAvailable": table["totalAvailable"],
        }
        if public != expected_public:
            raise ValueError(f"공개 집계가 원천 교차검사 값과 다릅니다: {table['name']}")

    ai_catalog = next(row for row in catalog if row[1] == AI_YOUTH_CLUB)
    if ai_catalog[3] not in (21, 22):
        raise ValueError(f"개설 현황의 {AI_YOUTH_CLUB} 정원 진단값이 예상 범위를 벗어났습니다.")
    ai_public = next((club for club in public_clubs if club["name"] == AI_YOUTH_CLUB), None)
    if ai_public and ai_public["capacity"] != 22:
        raise ValueError(f"{AI_YOUTH_CLUB}의 공개 운영 정원은 22명이어야 합니다.")

    public_strings = {value for value in public_string_values(data) if value}
    identities = exact_identity_values(teacher_names + student_names)
    if identities & public_strings:
        raise ValueError("원천 학생·교사 이름이 공개 JS의 문자열 값으로 포함되었습니다.")
    if set(narrative_values) & public_strings:
        raise ValueError("원천 내부 서술이 공개 JS의 문자열 값으로 포함되었습니다.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("snapshot", type=Path, help="공개 club_stats_2026_2.js")
    parser.add_argument("--workbook", type=Path, help="지원 현황 기준본 .xlsx")
    parser.add_argument("--catalog", type=Path, help="공식 개설 현황 .xlsx")
    parser.add_argument(
        "--expect-current-baseline",
        action="store_true",
        help="이번 배포의 2026-09-01·18/94/80/174 회귀값도 고정 검증",
    )
    args = parser.parse_args()
    if bool(args.workbook) != bool(args.catalog):
        parser.error("--workbook과 --catalog는 함께 지정해야 합니다.")

    data, raw = load_public(args.snapshot)
    validate_public(data, raw, args.expect_current_baseline)
    if args.workbook and args.catalog:
        validate_sources(data, args.workbook, args.catalog)

    totals = data["totals"]
    source_message = " · 원천 4방향 교차검사" if args.workbook else ""
    print(
        f"공개 통계 검증 PASS: {totals['clubs']}개 동아리, "
        f"1학년 {totals['grade1']['available']}명, "
        f"2학년 {totals['grade2']['available']}명, 전체 {totals['available']}명, "
        f"개인정보·내부 링크 0건{source_message}"
    )


if __name__ == "__main__":
    main()
